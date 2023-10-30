import abc
from pathlib import Path

import openpyxl

import config
from domain.models import Route


class AbstractRouteRepository(abc.ABC):
    """Абстрактный репозиторий маршрутов. Задает фреймворк и основные методы
    для работы с хранилищем данных. Конкретные реализации репозитория
    должны наследоваться от этого класса."""

    def add_all(self, routes: list[Route]) -> None:
        """Метод для добавления списка маршрутов в репозиторий."""
        self._add_all(routes)

    @abc.abstractmethod
    def _add_all(self, routes: list[Route]) -> None:
        """Абстрактный метод добавления данных маршрута в репозиторий."""
        raise NotImplementedError


class ExcelRouteRepository(AbstractRouteRepository):
    """Класс репозитория маршрутов, выгружающий результаты в excel-файл."""
    def __init__(
        self,
        weight_list: list[int | float],
        mode: str,
        routes_to_parse: list[tuple],
        results_db_name: Path = config.RESULTS_DB_NAME,
    ) -> None:
        super().__init__()
        self.weight_list = weight_list
        self.results_db_name = results_db_name
        self.results_db_name.touch()
        self.mode = mode
        self.last_column = 0
        self.routes_to_parse = routes_to_parse

    def _add_all(self, routes: list[Route]) -> None:
        """Метод для сохранения результатов парсинга. Он
        создает книгу Excel наполняет её данными и сохраняет."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'results'
        self._make_headers(worksheet)

        row_titles = sorted(self.routes_to_parse)
        self._make_row_titles(row_titles, worksheet)
        self._fill_data(routes, row_titles, worksheet)

        workbook.save(self.results_db_name)
        workbook.close()

    def _make_headers(self, worksheet):
        """Вспомогательный метод для создания шапки в Excel-файле."""

        worksheet.append(
            ('Город (Отправитель)', 'Город (Получатель)',
             'СДЭК', f'{self.mode}'))
        string_two = (('', '', 'Вес, кг')
                      + ('',) * (len(self.weight_list) - 1)
                      + ('Срок, дней',))
        string_three = ('', '') + tuple(self.weight_list) + ('от', 'до')
        self.last_column = len(string_three)
        worksheet.append(string_two)
        worksheet.append(string_three)

        self._make_cells_bold(
            worksheet,
            rows=(1, 3),
            columns=(1, self.last_column))
        self._adjust_column_width(worksheet, row=1, columns=(1, 2))
        self._adjust_column_width(worksheet, row=3,
                                  columns=(3, self.last_column))

    def _get_row_titles(self, routes):
        """Вспомогательный метод для вытаскивания пунктов отправки и назначения
        из переданных данных."""
        row_titles = set((item.city_from, item.city_to) for item in routes)
        return sorted(row_titles)

    def _make_row_titles(self, row_titles, worksheet):
        """Вспомогательный метод для задания названий строк."""
        for city_from, city_to in row_titles:
            worksheet.append((city_from, city_to))

    def _fill_data(
        self,
        routes: list,
        row_titles: list,
        worksheet: openpyxl.worksheet.worksheet.Worksheet
    ) -> None:
        """Вспомогательный метод для сохрания спарсенных значений
        в подготовленную таблицу."""
        for item in routes:
            row = 4 + row_titles.index((item.city_from, item.city_to))
            column = 3 + self.weight_list.index(item.weight)
            worksheet.cell(row, column).value = item.total_cost
            worksheet.cell(row,
                           self.last_column - 1).value = item.duration_min
            worksheet.cell(row, self.last_column).value = item.duration_max

    @staticmethod
    def _make_cells_bold(
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        rows: list[int] | tuple[int],
        columns: list[int] | tuple[int],
    ) -> None:
        """Вспомогательный метод для выделения диапазона ячеек на
        листе Excel-книги жирным шрифтом."""
        bold_font = openpyxl.styles.Font(bold=True)
        for row in range(rows[0], rows[1] + 1):
            for column in range(columns[0], columns[1] + 1):
                worksheet.cell(row, column).font = bold_font

    @staticmethod
    def _adjust_column_width(
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        row: int,
        columns: list[int] | tuple[int],
    ) -> None:
        """Вспомогательный метод для изменения ширины ячеек у заданного
        диапазона на листе Excel-книги."""
        for column in range(columns[0], columns[1] + 1):
            column_letter = worksheet.cell(row, column).column_letter
            column_width = len(str(worksheet.cell(row, column).value)) + 7
            worksheet.column_dimensions[column_letter].width = column_width
